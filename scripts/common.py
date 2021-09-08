"""
Common functionalities between scripts.
"""

import pandas as pd

from datetime import datetime
from time import time
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def set_items_df_column_order(items_df):
    """
    Set the desired order of columns to write to an excel spreadsheet on items dataframes

    :param items_df: items dataframe to have its column order set

    :returns: new df with the desired column order
    """
    return items_df[[
        "app_id",
        "name",
        "price_unitary",
        "amount",
        "price_total",
        "api_error",
        "price_date",
        "price_date_timestamp",
        "market_hash_name",
    ]]


def set_summary_df_column_order(summary_df):
    """
    Set the desired order of columns to write to an excel spreadsheet on summary dataframes

    :param summary_df: summary dataframe to have its column order set

    :returns: new df with the desired column order
    """
    return summary_df[["price_date", "price_total", "api_error"]]


def format_workbook(workbook):
    """
    Format workboot with sheets for each day and one for summary

    :param workbook: openpyxl workbook

    :returns: nothing
    """
    # set column common styles
    column_font = Font(size = 12, name = "Arial")
    column_alignment = Alignment(horizontal="center", vertical="center")

    # set summary worksheet column styles
    # format: ["name", "width", "font", "alignment"]
    summary_columns = [
        ["A", 14, column_font, column_alignment],
        ["B", 14, column_font, column_alignment],
        ["C", 14, column_font, column_alignment],
    ]

    # set any date worksheet column styles
    # format: ["name", "width", "font", "alignment"]
    date_columns = [
        ["A", 12, column_font, column_alignment],
        ["B", 55, column_font, column_alignment],
        ["C", 15, column_font, column_alignment],
        ["D", 10, column_font, column_alignment],
        ["E", 15, column_font, column_alignment],
        ["F", 11, column_font, column_alignment],
        ["G", 14, column_font, column_alignment],
        ["H", 24, column_font, column_alignment],
        ["I", 55, column_font, column_alignment],
    ]

    # TODO
    # header font name and size
    # price columns formatting
    for worksheet in workbook.worksheets:

        # set which columns to style
        columns = summary_columns if worksheet.title == "Summary" else date_columns

        # style columns
        for name, width, font, alignment in columns:
            worksheet.column_dimensions[name].font = font
            worksheet.column_dimensions[name].alignment = alignment
            worksheet.column_dimensions[name].width = width

        # format header row
        worksheet.row_dimensions[1] = Font(size = 12, name = "Arial", bold = True)


def write_items_to_excel(items, summary, excel_file_name):
    """
    Write an array of items dict into a dataframe.
    Also, compute the sum value of all items.

    :param items: list of dict containing items
    :param summary: list of dict with prices sum by date
    :param excel_file_name: file name to write to

    :returns: nothing
    """
    # get today date
    today_date = datetime.utcnow().strftime("%Y-%m-%d")

    # NOTE: Excel Writter must be created AFTER workbook is loaded!
    #
    # retrieve existent sheets in excel_file_name
    try:
        excel_book = load_workbook(excel_file_name)

    # error -> no file name in current directory: do nothing
    except:
        excel_writer = pd.ExcelWriter(excel_file_name, engine="openpyxl")
        has_summary_spreadsheet = False
        print("Excel file name provided not found. Initializing empty one.")

    # success -> keep already present sheets, and remove the ones that will be overwritten
    else:
        has_summary_spreadsheet = False
        if "Summary" in excel_book.sheetnames:
            has_summary_spreadsheet = True
            excel_book.remove(excel_book["Summary"])
        if today_date in excel_book.sheetnames:
            excel_book.remove(excel_book[today_date])
        excel_writer = pd.ExcelWriter(excel_file_name, engine="openpyxl")
        excel_writer.book = excel_book

    # set items as a dataframe and compute total price of each iten
    items_today_df = pd.DataFrame(items)
    items_today_df["price_total"] = items_today_df["price_unitary"] * items_today_df["amount"]

    # add today's prices sum to dataframe
    today_price_total = items_today_df["price_total"].sum()
    api_error_amount = items_today_df[items_today_df["api_error"] == "yes"].shape[0]
    today_sum = {
        "amount": items_today_df["amount"].sum(),
        "app_id": "---",
        "market_hash_name": "---",
        "name": "Sum of all items",
        "price_unitary": "---",
        "price_total": today_price_total,
        "price_date": today_date,
        "price_date_timestamp": int(time()),
        "api_error": "yes" if api_error_amount > 0 else "no"
    }
    items_today_df = items_today_df.append(today_sum, ignore_index=True)

    # write today's prices to excel exclusive sheet
    items_today_df = set_items_df_column_order(items_today_df)
    items_today_df.to_excel(excel_writer, index=False, sheet_name=today_date)

    # add today's summary (update current row date if it exists already)
    summary_df = pd.DataFrame(summary)
    today_summary = {
        "api_error": today_sum["api_error"],
        "price_total": today_sum["price_total"],
        "price_date": today_sum["price_date"],
    }
    last_summary_line = summary_df.shape[0] - 1
    if has_summary_spreadsheet and summary_df["price_date"][last_summary_line] == today_date:
        summary_df.at[last_summary_line, "api_error"] = today_summary["api_error"]
        summary_df.at[last_summary_line, "price_total"] = today_summary["price_total"]
    else:
        summary_df = summary_df.append(today_summary, ignore_index=True)

    # write today's prices summary to summary sheet
    summary_df = set_summary_df_column_order(summary_df)
    summary_df.to_excel(excel_writer, index=False, sheet_name='Summary')

    # persis changes
    format_workbook(excel_writer.book)
    excel_writer.save()
